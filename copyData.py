"""
from openpyxl.utils import rows_from_range

def copy_range(range_str, src, dst):

    for row in rows_from_range(range_str):
        for cell in row:
            dst[cell].value = src[cell].value

    return

"""



from openpyxl import load_workbook


def copyColumn(workbook):
    worksheet = workbook.worksheets[0]
    worksheet1 = workbook.worksheets[1]

    for row in range(1, worksheet.max_row):
        col = worksheet.cell(row=row, column=2)
        worksheet1.cell(row=row, column=6, value=col.value)

    workbook.save('Data.xlsx')

if __name__ == '__main__':
    copyColumn(load_workbook('Data.xlsx'))