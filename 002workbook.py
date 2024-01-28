from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def holidays(workbook, holidates):
    worksheet = workbook.worksheets[1]
    # Форматируем праздники в формат datetime
    dates = [datetime.strptime(date, '%d.%m.%Y').date() for date in holidates]
    # Итерируем и находим колонку с нужными нам значениями
    for row in range(1, worksheet.max_row +1):
        days = worksheet.cell(row=row, column=1).value
        # Проверяем даты на соответствие примеру
        if days and isinstance(days, datetime):
            days_dates = days.date()
            # Проверяем даты на соответствие праздничных дат
            if days_dates in dates:
                for cell in worksheet[row]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    workbook.save("Colorexcel.xlsm")

if __name__ == "__main__":
    holidays(
        load_workbook("Delovniki - DECEMBER 2023.xlsm"), 
        [
        '01.01.2023', '02.01.2023', '08.02.2023', 
        '27.04.2023', '01.05.2023', '08.06.2023', 
        '25.06.2023', '14.08.2023', '17.08.2023', 
        '15.09.2023', '23.09.2023', '25.10.2023', 
        '31.10.2023', '01.11.2023', '23.11.2023', 
        '25.12.2023', '26.12.2023'
        ]
        )