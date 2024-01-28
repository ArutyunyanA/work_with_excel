from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

def total_line(workbook):
    # Устанавливаем рабочий лист для работы
    worksheet = workbook.worksheets[1]
    # Находим первую строку с днем недели
    for row in range(1, worksheet.max_row + 1):
        week_days = worksheet.cell(row=row, column=2).value
        # Если в списке будет 'nedelja' то вставляем строку после этого поля
        if week_days == 'nedelja':
            worksheet.insert_rows(row + 1, amount=1)
            # Записываем "Skupaj" в новую ячейку
            worksheet.cell(row=row + 1, column=1, value="Skupaj").font = Font(bold=True)
            # Устанавливаем рамку для новой строки
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            for cell in worksheet[row + 1]:
                cell.border = border
    # Сохраняем excel file        
    workbook.save('newfile.xlsm')

if __name__ == '__main__':
    total_line(
        load_workbook('Delovniki - DECEMBER 2023.xlsm')
    )
    
    
    