from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

def lunch_time(workbook):
    # Определяем рабочую таблицу
    worksheet = workbook.worksheets[1]
    for row in range(1, worksheet.max_row):
        interval = str(worksheet.cell(row=row, column=5).value)
        # Проеверяем наличие символа '-' в строке
        if '-' in interval:
            # Разбиваем строку на начальное и конечное время
            start_time, end_time = interval.split('-')
            try:
                # Перобразуем строки в формат времени
                start_time = datetime.strptime(start_time, "%H:%M")
                end_time = datetime.strptime(end_time, "%H:%M")
                # Вычисляем разницу
                time_delta = end_time - start_time
                # Делаем проверку
                if time_delta > timedelta(minutes=30):
                    worksheet.cell(row=row, column=5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    workbook.save("NewBook.xlsm")

                # Выводим результат
                # !!! Нужно понимать куда вставлять данные и что с ними делать !!!
                print(f"Pavza malica {row}: {time_delta}")
            except ValueError as error:
                print(f"No values Error: {row}: {error}")
        else:
            print(f"No values Error: {row}")
        

if __name__ == '__main__':
    lunch_time(load_workbook('Delovniki - DECEMBER 2023.xlsm'))
